#coding:utf-8
import urllib.request
import gzip
from bs4 import BeautifulSoup
import datetime
import os
import openpyxl


today1 = datetime.date.today()
yesterday1 = today1 - datetime.timedelta(days=1)
today2 = str(today1)
yesterday = str(yesterday1)


url = "http://caipiao.163.com/award/cqssc/"+today2[0:4]+today2[5:7]+today2[8:10]+".html"
request = urllib.request.Request(url)
response = urllib.request.urlopen(request)
data = response.read()
data=gzip.decompress(data)
data = data.decode('utf-8','ignore')
#print(data)
soup = BeautifulSoup(data,"html.parser")
tables=soup.findAll('table')
tab=tables[0]
dict={}
for tr in tab.findAll('tr'):
    for td in tr.findAll('td'):
        #print(td.get('data-period'))
        if dict.__contains__(td.get('data-period')):
            
            #dict.get(td.get('data-period'))
            continue
        else:
            dict[td.get('data-period')]=td.get('data-win-number')
        #print(td.get('data-win-number'))
        #print(d['180429001'])
#for i in range(1,121):
    #print(dict[str(180429000+i)])

url = "http://caipiao.163.com/award/cqssc/"+yesterday[0:4]+yesterday[5:7]+yesterday[8:10]+".html"
request = urllib.request.Request(url)
response = urllib.request.urlopen(request)
data = response.read()
data=gzip.decompress(data)
data = data.decode('utf-8','ignore')
#print(data)
soup = BeautifulSoup(data,"html.parser")
tables=soup.findAll('table')
tab=tables[0]
for tr in tab.findAll('tr'):
    for td in tr.findAll('td'):
        #print(td.get('data-period'))
        if dict.__contains__(td.get('data-period')):
            
            #dict.get(td.get('data-period'))
            continue
        else:
            dict[td.get('data-period')]=td.get('data-win-number')
            
            
dict.pop(None)
dict={key: value for key, value in dict.items() if value !=None}

#filename='301-2018-' + yesterday[5:7] + '-' + yesterday[8:10] + '-2018-'+ today2[5:7]+ '-' + today2[8:10]+'.txt'

#f=open(filename,"w")
#for i in dict:
    #print (i,dict[i])
    #f.write(str(i) + '                 ' + str(dict[i]) + '\n')
#f.close()
#print(os.getcwd())

wb=openpyxl.Workbook()
ws=wb.active
r=1
for i in dict:
    ws.cell(row=r,column=1).value=str(i)
    ws.cell(row=r,column=2).value=str(dict[i])
    r=r+1
wb.save('原始数据'+yesterday[0:4]+yesterday[5:7]+yesterday[8:10]+'-'+today2[0:4]+today2[5:7]+today2[8:10]+'.xlsx')
