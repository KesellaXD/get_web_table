#coding:utf-8
import urllib.request
import gzip
from bs4 import BeautifulSoup
import datetime
import os
import openpyxl
import time


def download():
    day1=input('请输入起始日期，格式为yyyymmdd(如20180401）:')
    print('\n')
    day2=input('请输入终止日期，格式为yyyymmdd(如20180402）:')
    print('\n')

    try:
        day1=datetime.datetime(int(day1[0:4]),int(day1[4:6]),int(day1\
        [6:8]))
        day2=datetime.datetime(int(day2[0:4]),int(day2[4:6]),int(day2\
        [6:8]))
        today=str(datetime.date.today())
        today=datetime.datetime(int(today[0:4]),int(today[5:7]),int\
        (today[8:10]))
    except:
        print('请输入正确的日期!\n')
        download()

    if((today<day1) or (today<day2)):
        print('请输入正确的日期!\n')
        download()
        
    d=abs((day1-day2).days)
    if(day1>day2):
        day3=day1
        day1=day2
        day2=day3
    day1_=str(day1)
    day2_=str(day2)
    print('你选择的范围是'+day1_[0:4]+'年'+day1_[5:7]+'月'+day1_[8:10]+\
    '日到'+day2_[0:4]+'年'+day2_[5:7]+'月'+day2_[8:10]+'日!\n')
    time.sleep(1)
    print('开始下载...\n')
    dict={}
    for i in range(0,d+1):
        thatday1 = day2 - datetime.timedelta(days=i)
        thatday = str(thatday1)
        url = "http://caipiao.163.com/award/cqssc/"+thatday[0:4]+\
        thatday[5:7]+thatday[8:10]+".html"
        request = urllib.request.Request(url)
        response = urllib.request.urlopen(request)
        data = response.read()
        data=gzip.decompress(data)
        data = data.decode('utf-8','ignore')
        soup = BeautifulSoup(data,"html.parser")
        tables=soup.findAll('table')
        tab=tables[0]
        for tr in tab.findAll('tr'):
            for td in tr.findAll('td'):
                if dict.__contains__(td.get('data-period')):
                    continue
                else:
                    dict[td.get('data-period')]=td.get('data-win-number')
        print(thatday[0:4]+'年'+thatday[5:7]+'月'+thatday[8:10]+\
        '日的数据下载完成!\n')
        
    dict.pop(None)
    #dict={key: value for key, value in dict.items() if value !=None}
    day1=str(day1)
    day2=str(day2)
    wb=openpyxl.Workbook()
    ws=wb.active
    r=1
    for i in dict:
        ws.cell(row=r,column=1).value=str(i)
        if str(dict[i])=="None":
            ws.cell(row=r,column=2).value=""
        else:
            ws.cell(row=r,column=2).value=str(dict[i])
        r=r+1
    wb.save(day1[0:4]+'-'+day1[5:7]+'-'+day1[8:10]+'-'+day2[0:4]+'-'+\
    day2[5:7]+'-'+day2[8:10]+'重庆时时彩开奖数据.xlsx')
    
start=time.clock()
download()
end=time.clock()
print('所有数据下载完成，耗时%s秒!\n'%(end-start))
input('按回车键退出程序...')
